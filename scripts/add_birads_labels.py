"""
Modify nnDetection JSON annotation files to add BI-RADS multi-class labels.

Class mapping:
  BI-RADS 2     -> class 0
  BI-RADS 3     -> class 1
  BI-RADS 4/4a/4b/4c -> class 2

Sources:
  - Test set: per-lesion BI-RADS from Excel (病灶标注对应BI-rads分类.xlsx)
  - Training set: folder-level BI-RADS from dataset_statistics.csv (birads2/3/4)

Also updates dataset.json to reflect 3 classes.
"""

import csv
import json
import os
import shutil
from pathlib import Path

import openpyxl


def parse_birads(val):
    """Convert BI-RADS value to class ID: 2->0, 3->1, 4/4a/4b/4c->2"""
    s = str(val).strip().lower()
    if s == '2':
        return 0
    elif s == '3':
        return 1
    elif s.startswith('4'):
        return 2
    else:
        raise ValueError(f"Unknown BI-RADS value: {val}")


def load_test_birads(excel_path):
    """
    Parse per-lesion BI-RADS from Excel.
    Returns: {filename: {inst_id(int): birads_class(int)}}
    """
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    result = {}
    current_file = None

    for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
        b_val = row[1].value  # column B = filename
        c_val = row[2].value  # column C = lesion id (e.g. "1-红", "2-绿")
        d_val = row[3].value  # column D = BI-RADS

        if b_val is not None:
            current_file = str(b_val).strip()
            # Ensure .nii extension
            if not current_file.endswith('.nii'):
                current_file += '.nii'
            if current_file not in result:
                result[current_file] = {}

        if c_val is not None and d_val is not None and current_file is not None:
            # Extract instance number from "1-红", "2-绿", or just "1"
            inst_str = str(c_val).strip()
            inst_num = int(inst_str.split('-')[0])
            birads_class = parse_birads(d_val)
            result[current_file][inst_num] = birads_class

    return result


def main():
    data_dir = Path("/Volumes/Lang/Research/Data/3D Ultrasound/nnDet/Duying")
    excel_path = Path("/Volumes/Autzoko/Dataset/Ultrasound/已标注及BI-rads分类20260123/病灶标注对应BI-rads分类.xlsx")
    csv_path = data_dir / "dataset_statistics.csv"

    labels_tr = data_dir / "raw_splitted" / "labelsTr"
    labels_ts = data_dir / "raw_splitted" / "labelsTs"

    # --- Step 0: Backup ---
    for d in [labels_tr, labels_ts]:
        backup = d.parent / (d.name + "_backup_binary")
        if not backup.exists():
            print(f"Backing up {d.name} -> {backup.name}")
            shutil.copytree(d, backup)
        else:
            print(f"Backup already exists: {backup.name}")

    dataset_json = data_dir / "dataset.json"
    dataset_json_backup = data_dir / "dataset_backup_binary.json"
    if not dataset_json_backup.exists():
        shutil.copy2(dataset_json, dataset_json_backup)
        print("Backed up dataset.json")

    # --- Step 1: Load test set per-lesion BI-RADS ---
    print("\nParsing test set BI-RADS from Excel...")
    test_birads = load_test_birads(excel_path)
    print(f"  Found {len(test_birads)} test files with per-lesion BI-RADS")
    for fname, mapping in sorted(test_birads.items()):
        print(f"    {fname}: {mapping}")

    # --- Step 2: Load CSV for case_id <-> filename mapping ---
    with open(csv_path) as f:
        reader = csv.DictReader(f)
        cases = list(reader)

    # --- Step 3: Modify JSON files ---
    modified_tr = 0
    modified_ts = 0
    errors = []

    for case in cases:
        cid = case['volume_id']
        split = case['split']
        lesion_class = case['lesion_class']
        fname = case['image_path'].split('/')[-1]

        if split == 'test':
            json_path = labels_ts / f"{cid}.json"
        else:
            json_path = labels_tr / f"{cid}.json"

        if not json_path.exists():
            continue

        with open(json_path) as f:
            data = json.load(f)

        instances = data.get('instances', {})
        new_instances = {}

        if split == 'test':
            # Use per-lesion BI-RADS from Excel
            # Match filename (without .ai extension)
            fname_nii = fname.replace('.ai', '.nii') if fname.endswith('.ai') else fname
            if fname_nii not in test_birads:
                # Try without extension variations
                fname_base = fname_nii.replace('.nii', '')
                found = False
                for k in test_birads:
                    if k.replace('.nii', '') == fname_base:
                        fname_nii = k
                        found = True
                        break
                if not found:
                    errors.append(f"Test case {cid} ({fname}) not found in Excel")
                    # Fallback: keep as class 0
                    new_instances = instances
                    data['instances'] = new_instances
                    with open(json_path, 'w') as f:
                        json.dump(data, f, indent=2)
                    continue

            birads_map = test_birads[fname_nii]
            for inst_id_str, old_class in instances.items():
                inst_num = int(inst_id_str)
                if inst_num in birads_map:
                    new_instances[inst_id_str] = birads_map[inst_num]
                else:
                    errors.append(f"Test {cid} ({fname}) inst {inst_num} not in Excel")
                    new_instances[inst_id_str] = old_class
            modified_ts += 1
        else:
            # Training set: use folder-level class
            birads_class = parse_birads(
                '2' if lesion_class == 'birads2' else
                '3' if lesion_class == 'birads3' else
                '4' if lesion_class == 'birads4' else
                lesion_class
            )
            for inst_id_str in instances:
                new_instances[inst_id_str] = birads_class
            modified_tr += 1

        data['instances'] = new_instances
        with open(json_path, 'w') as f:
            json.dump(data, f, indent=2)

    print(f"\nModified {modified_tr} training JSONs, {modified_ts} test JSONs")

    if errors:
        print(f"\nWarnings ({len(errors)}):")
        for e in errors:
            print(f"  - {e}")

    # --- Step 4: Update dataset.json ---
    print("\nUpdating dataset.json...")
    with open(dataset_json) as f:
        ds = json.load(f)

    ds['labels'] = {
        "0": "BIRADS2",
        "1": "BIRADS3",
        "2": "BIRADS4"
    }

    with open(dataset_json, 'w') as f:
        json.dump(ds, f, indent=2)
    print("  labels updated to:", ds['labels'])

    # --- Step 5: Verify a few cases ---
    print("\nVerification (sample cases):")
    for cid in ['case_00000', 'case_00003', 'case_00013', 'case_00057']:
        for label_dir in [labels_ts, labels_tr]:
            jp = label_dir / f"{cid}.json"
            if jp.exists():
                with open(jp) as f:
                    d = json.load(f)
                src = "Ts" if "Ts" in str(label_dir) else "Tr"
                matching_case = [c for c in cases if c['volume_id'] == cid][0]
                fname = matching_case['image_path'].split('/')[-1]
                print(f"  {cid} ({fname}, {src}): {d['instances']}")

    print("\nDone! You need to re-run preprocessing (nndet_prep) for these changes to take effect.")


if __name__ == '__main__':
    main()
